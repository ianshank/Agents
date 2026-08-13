"""Workbook writer for the generated end-to-end test matrix.

Kept separate from :mod:`tests._e2e_matrix` for two reasons. It is the only part of the
matrix that needs a third-party package, so isolating it keeps the engine importable with
zero dependencies and the CSV/markdown artifacts reachable on a machine that has no
spreadsheet library. And it is the only part with a reproducibility hazard worth naming:

``openpyxl`` stamps ``dcterms:created``/``modified`` with the wall clock, and ``zipfile``
stamps every archive entry with the local time at write. An ``.xlsx`` is a zip, so a
workbook saved twice from identical data differs in bytes both ways. Since this artifact
is committed and freshness-gated, both stamps are pinned to a value derived from the run
being reported, never from "now" - see :func:`write_workbook`.
"""

from __future__ import annotations

import datetime as dt
import io
import logging
import re
import zipfile
from collections.abc import Sequence
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from tests._e2e_matrix import (
    NOT_RUN,
    STATUS_COLUMN,
    STATUS_FAIL,
    STATUS_PASS,
    STATUS_SKIP,
    Sheet,
    safe_cell,
)

logger = logging.getLogger(__name__)

#: Named so the caller gets an actionable message instead of a bare ModuleNotFoundError.
INSTALL_HINT = "Install with: pip install 'langfuse-eval-harness[e2e-matrix]'"

#: Earliest timestamp the ZIP format can represent; the floor for a pinned stamp.
ZIP_EPOCH = (1980, 1, 1, 0, 0, 0)

#: Latest timestamp the ZIP format can represent: its DOS date packs the year into 7 bits
#: counting from 1980. Without a ceiling a far-future stamp raises deep inside `zipfile`.
ZIP_MAX_YEAR = 2107

#: ZIP stores times in MS-DOS format, which encodes seconds as seconds/2 - odd seconds are
#: not representable and get truncated on write. Flooring to even here means the value this
#: function returns is exactly the value the archive will carry, so a caller can compare the
#: two without knowing about the format's granularity.
ZIP_SECOND_GRANULARITY = 2


@dataclass(frozen=True)
class WorkbookStyle:
    """Presentation knobs. A dataclass rather than literals at call sites (AGENTS.md).

    Widths are in Excel character units, which is what ``column_dimensions[...].width``
    expects; the min/max bracket keeps a long ``Detail`` cell from making a column
    unreadable while still letting short columns shrink.
    """

    min_column_width: int = 10
    max_column_width: int = 60
    width_padding: int = 2
    header_row: int = 1
    first_data_row: int = 2
    header_fill: str = "FF2F3E4E"
    header_font: str = "FFFFFFFF"
    #: Fill per status. Every key comes from the engine rather than being respelled here: a
    #: rename there would otherwise leave this table silently matching nothing.
    status_fills: tuple[tuple[str, str], ...] = (
        (STATUS_PASS, "FFD5EFD8"),
        (STATUS_FAIL, "FFF6CFCF"),
        (STATUS_SKIP, "FFFBEFCB"),
        (NOT_RUN, "FFE6E6E6"),
    )


DEFAULT_STYLE = WorkbookStyle()


def _require_openpyxl() -> Any:
    """Import openpyxl, or explain how to get it.

    Mirrors the ParquetDataset seam: the dependency is optional at install time, and the
    one code path that needs it fails with an actionable message rather than a traceback
    from an unrelated import line.
    """
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - exercised via sys.modules injection
        raise ImportError(f"openpyxl is required to write the .xlsx matrix. {INSTALL_HINT}") from exc
    return openpyxl


def parse_stamp(iso_timestamp: str) -> tuple[int, int, int, int, int, int]:
    """An ISO-8601 timestamp as the six UTC components ``zipfile`` wants.

    An offset-bearing input is converted to UTC first. The components this returns are
    written into ``docProps/core.xml`` with a ``Z`` suffix, so taking them from a
    ``-07:00`` stamp verbatim would relabel local time as UTC and move the instant by the
    offset. That is not hypothetical: the provenance stamp comes from ``git log --format=%cI``,
    which carries the committer's offset, so it is UTC only by accident of where it ran.
    A naive stamp is treated as UTC, which is what the callers document it to be.

    Falls back to the ZIP epoch rather than the current time: a malformed stamp must not
    silently reintroduce the non-determinism this function exists to remove.
    """
    try:
        parsed = dt.datetime.fromisoformat(iso_timestamp.replace("Z", "+00:00"))
    except ValueError:
        logger.warning("unparseable provenance timestamp %r; pinning the archive to the ZIP epoch", iso_timestamp)
        return ZIP_EPOCH
    if parsed.tzinfo is not None:
        parsed = parsed.astimezone(dt.timezone.utc).replace(tzinfo=None)
    if parsed.year < ZIP_EPOCH[0] or parsed.year > ZIP_MAX_YEAR:
        logger.warning("provenance timestamp %r is outside the ZIP range; pinning to the epoch", iso_timestamp)
        return ZIP_EPOCH
    second = parsed.second - (parsed.second % ZIP_SECOND_GRANULARITY)
    return (parsed.year, parsed.month, parsed.day, parsed.hour, parsed.minute, second)


#: The OPC part carrying document metadata, and the element openpyxl stamps at save time.
CORE_PROPERTIES_PART = "docProps/core.xml"
_MODIFIED_RE = re.compile(rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)")


def _pin_modified(raw: bytes, stamp: tuple[int, int, int, int, int, int]) -> bytes:
    """Rewrite ``dcterms:modified`` in ``docProps/core.xml`` to the pinned stamp.

    Setting ``workbook.properties.modified`` before saving does not survive: openpyxl
    overwrites it with the wall clock as it writes. ``created`` *is* respected, so a workbook
    saved twice from identical data differed in exactly one element -- enough to defeat the
    byte comparison the freshness gate relies on. Patching the written part is the only
    place the value can be pinned.
    """
    replacement = "{:04d}-{:02d}-{:02d}T{:02d}:{:02d}:{:02d}Z".format(*stamp).encode()
    return _MODIFIED_RE.sub(rb"\g<1>" + replacement + rb"\g<2>", raw)


def _normalise_archive(raw: bytes, stamp: tuple[int, int, int, int, int, int]) -> bytes:
    """Rewrite every zip entry with a pinned timestamp, preserving order and content.

    Entry order is preserved rather than sorted: the OPC package puts ``[Content_Types].xml``
    first by convention, and openpyxl already emits a deterministic order for identical
    input, so reordering would buy nothing and risk reader compatibility.
    """
    buffer = io.BytesIO()
    with (
        zipfile.ZipFile(io.BytesIO(raw)) as source,
        zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as target,
    ):
        for name in source.namelist():
            info = zipfile.ZipInfo(name, date_time=stamp)
            info.compress_type = zipfile.ZIP_DEFLATED
            payload = source.read(name)
            if name == CORE_PROPERTIES_PART:
                payload = _pin_modified(payload, stamp)
            target.writestr(info, payload)
    return buffer.getvalue()


def _column_width(column: str, values: Sequence[str], style: WorkbookStyle) -> int:
    """Width that fits the longest cell, clamped to the configured bracket."""
    longest = max((len(value) for value in (column, *values)), default=style.min_column_width)
    return max(style.min_column_width, min(style.max_column_width, longest + style.width_padding))


def _write_sheet(worksheet: Any, sheet: Sheet, style: WorkbookStyle) -> None:
    """Populate one worksheet: header, rows, widths, freeze pane, filter, status colour."""
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    # Scrubbed once per cell and reused below for the append, the column-width measurement,
    # and the status-colour lookup -- previously each ran its own `safe_cell` pass over every
    # row, doing the same scrub up to three times per cell.
    safe_rows = [[safe_cell(cell) for cell in row] for row in sheet.rows]

    worksheet.append(list(sheet.columns))
    for row in safe_rows:
        worksheet.append(row)

    for cell in worksheet[style.header_row]:
        cell.font = Font(bold=True, color=style.header_font)
        cell.fill = PatternFill(fill_type="solid", start_color=style.header_fill)

    for index, column in enumerate(sheet.columns, start=1):
        values = [row[index - 1] for row in safe_rows]
        worksheet.column_dimensions[get_column_letter(index)].width = _column_width(column, values, style)

    if STATUS_COLUMN in sheet.columns:
        status_index = sheet.columns.index(STATUS_COLUMN)
        colours = dict(style.status_fills)
        letter = get_column_letter(status_index + 1)
        for offset, row in enumerate(safe_rows):
            colour = colours.get(row[status_index])
            if colour:
                worksheet[f"{letter}{style.first_data_row + offset}"].fill = PatternFill(
                    fill_type="solid", start_color=colour
                )

    worksheet.freeze_panes = f"A{style.first_data_row}"
    if sheet.rows:
        last = get_column_letter(len(sheet.columns))
        worksheet.auto_filter.ref = f"A{style.header_row}:{last}{len(sheet.rows) + 1}"


def write_workbook(
    sheets: Sequence[Sheet],
    path: Path,
    *,
    stamp_iso: str,
    creator: str = "tests/test_e2e_matrix.py",
    style: WorkbookStyle = DEFAULT_STYLE,
) -> Path:
    """Write the sheets to ``path`` as a byte-reproducible ``.xlsx``.

    ``stamp_iso`` is the run's provenance timestamp. It is written into the document
    properties and into every archive entry so that regenerating from unchanged inputs
    yields an identical file - without it the workbook would churn on every run and the
    freshness gate could never byte-compare it.
    """
    openpyxl = _require_openpyxl()

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    for sheet in sheets:
        _write_sheet(workbook.create_sheet(title=sheet.name), sheet, style)

    stamp = parse_stamp(stamp_iso)
    pinned = dt.datetime(*stamp)
    workbook.properties.created = pinned
    workbook.properties.modified = pinned
    workbook.properties.creator = creator
    workbook.properties.lastModifiedBy = creator

    raw = io.BytesIO()
    workbook.save(raw)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(_normalise_archive(raw.getvalue(), stamp))
    logger.info("wrote workbook %s (%d sheet(s))", path.as_posix(), len(sheets))
    return path
